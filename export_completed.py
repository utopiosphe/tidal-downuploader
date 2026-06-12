"""导出已完成任务到 Excel"""
import subprocess, csv

# 从远程 MySQL 导出数据
result = subprocess.run([
    "ssh", "root@117.55.199.29",
    "mysql --default-character-set=utf8mb4 -u tidal -ptidal_dl_2026 tidal_dl -N -e \""
    "SELECT id, track_id, title, artist, album, isrc, actual_quality, codec, "
    "file_size, s3_key, completed_at FROM tasks WHERE status='completed' ORDER BY id;\""
], capture_output=True)

raw = result.stdout.decode("utf-8", errors="replace")

rows = []
for line in raw.strip().split("\n"):
    if not line.strip():
        continue
    cols = line.split("\t")
    if len(cols) >= 10:
        s3_key = cols[9] if cols[9] != "NULL" else ""
        download_url = f"https://xiyaa.aybksd136.com/{s3_key}" if s3_key else ""
        try:
            file_size_mb = round(int(cols[8] or 0) / 1024 / 1024, 2)
        except:
            file_size_mb = 0
        rows.append({
            "ID": cols[0],
            "Track ID": cols[1],
            "标题": cols[2],
            "艺术家": cols[3],
            "专辑": cols[4],
            "ISRC": cols[5] if cols[5] != "NULL" else "",
            "音质": cols[6] if cols[6] != "NULL" else "",
            "编码": cols[7] if cols[7] != "NULL" else "",
            "文件大小(MB)": file_size_mb,
            "S3路径": s3_key,
            "下载链接": download_url,
            "完成时间": cols[10] if len(cols) > 10 and cols[10] != "NULL" else "",
        })

print(f"共 {len(rows)} 条完成任务")

# 写 Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "已完成任务"

    headers = list(rows[0].keys()) if rows else []
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
            cell.border = thin_border
            if h == "下载链接" and row[h]:
                cell.font = Font(color="0563C1", underline="single")

    # 自动列宽
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for r in rows[:50]:
            val_len = len(str(r.get(h, "")))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 60)

    output = "/Users/zt/Desktop/tidaltest/completed_tasks.xlsx"
    wb.save(output)
    print(f"✅ 已导出: {output}")

except ImportError:
    output = "/Users/zt/Desktop/tidaltest/completed_tasks.csv"
    with open(output, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"✅ 已导出 CSV: {output}")
